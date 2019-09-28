#-*- endcoding:utf-8
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from common.data.global_value import colorSingleton
class exceOperation(object):

    def __init__(self, file_path):
        self.file_path = file_path
        self.wb = load_workbook(file_path)
        self.colorSingleton = colorSingleton()

    def open_sheet(self, sheet_name):
        '''
        打开sheet页
        :param sheet_name: sheet页的名称
        :return:
        '''
        self.st = self.wb[sheet_name]
        return self.st

    def get_value(self, cellNum):
        '''
        获取cell的值
        '''
        return cellNum.value

    def get_cell_location(self, value):
        return "%s%d" % (value.coordinate, value.column)

    def set_cell_color(self, value):
        fill = PatternFill("solid", fgColor=self.colorSingleton.realityColor)
        value.fill = fill
        self.wb.save(self.file_path)

if __name__ == "__main__":
    fileP = r"D:\接口\Test_UI\data\lane\test.xlsx"
    sheetN = "Sheet1"
    ex = exceOperation(fileP)
    ex.get_datas(sheetN)
    reader = ex.get_readers()
    for i, j in reader.items():
        if i.value == "laneNum":
            ex.set_faile(i)
